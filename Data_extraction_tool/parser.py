import pandas as pd
ex_data = pd.read_excel('excel_file.xlsx')
ex_data['column_name'].value.toList()

def parser():
def type_of_file:
    ex_data = pd.read_excel('excel_file.xlsx')
    ex_csv = pd.read_csv('csv_file.csv')
'''
from openpyxl import load_workbook
wb = load_workbook('Book.xlsx')
ws = wb.active
for row in ws.iter_rows()
    for cell in row:
        print cell.valuecom
'''

from pdfquery import PDFQuery

pdf = PDFQuery('example.pdf')
pdf.load()

text_elements = pdf.pq('LTTextLineHorizontal')

text = [t.text for t in text_elements]

print(text)
'''
pdf = pdfquery.PDFQuery("tests/samples/IRS_1040A.pdf")
pdf.load()
label = pdf.pq('LTTextLineHorizontal:contains("Your first name and initial")')
left_corner = float(label.attr('x0'))
bottom_corner = float(label.attr('yo'))
name = pdf.pq('LTTextLineHorizontal:in_bbox("%s, %s, %s, %s")' % (left_corner, bottom_corner-30, left_corner+150, bottom_corner)).text()
name
'''
